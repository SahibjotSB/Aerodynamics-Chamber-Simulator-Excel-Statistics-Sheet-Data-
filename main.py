import pygame
import sys
import pymunk
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from datetime import datetime
import yagmail
import os
import time

# Initialize pygame and set up the display
def init():
    pygame.init()
    screen = pygame.display.set_mode((1200, 900))
    pygame.display.set_caption("Physics Simulation with Pymunk")
    return screen

# Create a dynamic circle
def create_circle(space, pos):
    body = pymunk.Body(1, 500, body_type=pymunk.Body.DYNAMIC)
    body.position = pos
    shape = pymunk.Circle(body, 7)
    shape.elasticity = 0.95
    shape.friction = 0.9
    space.add(body, shape)
    return shape

# Draw circles on the screen
def draw_circles(circles, screen, circle_surface):
    for circle in circles:
        pos_x = int(circle.body.position.x)
        pos_y = int(circle.body.position.y)
        circ_rect = circle_surface.get_rect(center=(pos_x, pos_y))
        screen.blit(circle_surface, circ_rect)

# Create static shapes (box, circle, polygon)
def create_shape(space, pos, shape_type):
    body = pymunk.Body(body_type=pymunk.Body.STATIC)
    body.position = pos
    if shape_type == 'box':
        shape = pymunk.Poly.create_box(body, (100, 50))
    elif shape_type == 'circle':
        shape = pymunk.Circle(body, 50)
    elif shape_type == 'polygon':
        vertices = [(-50, -30), (50, -30), (70, 0), (50, 30), (-50, 30), (-70, 0)]
        shape = pymunk.Poly(body, vertices)
    space.add(body, shape)
    return shape

# Draw static shapes on the screen
def draw_shapes(shapes, screen):
    for shape in shapes:
        if isinstance(shape, pymunk.Poly):
            vertices = shape.get_vertices()
            points = [(int(v.rotated(shape.body.angle).x + shape.body.position.x), int(v.rotated(shape.body.angle).y + shape.body.position.y)) for v in vertices]
            pygame.draw.polygon(screen, (217, 98, 50), points)
        elif isinstance(shape, pymunk.Circle):
            pos_x = int(shape.body.position.x)
            pos_y = int(shape.body.position.y)
            pygame.draw.circle(screen, (0, 0, 0), (pos_x, pos_y), int(shape.radius))

# Calculate the average speed of dynamic circles
def calculate_average_speed(circles):
    if not circles:
        return 0.0
    total_speed = sum(circle.body.velocity.length for circle in circles if circle.body.position.x < 1200)
    count = sum(1 for circle in circles if circle.body.position.x < 1200)
    return total_speed / count if count > 0 else 0.0

# Display the average speed at the top of the screen
def display_average_speed(average_speed, screen, font):
    speed_text = font.render(f"Average Speed: {average_speed:.2f}", True, (0, 0, 0))
    screen.blit(speed_text, (10, 10))

# Apply a simple drag force to the dynamic circles
def apply_aerodynamics(circles):
    drag_coefficient = 0.1
    for circle in circles:
        drag_force = -drag_coefficient * circle.body.velocity
        circle.body.apply_force_at_local_point(drag_force)

# Save data to Excel
def save_to_excel(data, filename='simulation_data.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Simulation Data"
    
    ws.append(["Time (s)", "Average Speed"])
    for entry in data:
        ws.append(entry)
    
    # Create the graph
    times, speeds = zip(*data)
    plt.figure(figsize=(10, 6))
    plt.plot(times, speeds, marker='o', linestyle='-', color='b')
    plt.title('Average Speed Over Time')
    plt.xlabel('Time (s)')
    plt.ylabel('Average Speed')
    plt.grid(True)
    graph_image_path = "speed_graph.png"
    plt.savefig(graph_image_path)
    plt.close()

    # Insert the graph into the Excel sheet
    img = XLImage(graph_image_path)
    img.anchor = f"{get_column_letter(ws.max_column + 2)}1"
    ws.add_image(img)
    
    wb.save(filename)
    os.remove(graph_image_path)

# Main function to run the simulation
def main():
    screen = init()
    clock = pygame.time.Clock()
    space = pymunk.Space()
    space.gravity = (500, 0)  # Horizontal gravity

    # Load and scale down the circle surface
    circle_surface = pygame.image.load('blue-glossy-ball-png.webp')
    circle_surface = pygame.transform.scale(circle_surface, (20, 20))  # Scale down to 20x20 pixels

    font = pygame.font.Font(None, 36)

    circles = []
    shapes = []
    shapes.append(create_shape(space, (900, 200), 'box'))
    shapes.append(create_shape(space, (900, 400), 'circle'))
    shapes.append(create_shape(space, (900, 600), 'polygon'))

    # Flag to track if the mouse button is pressed
    mouse_down = False

    # Data collection variables
    start_time = time.time()
    data_collection_interval = 5  # seconds
    last_data_collection_time = start_time
    data = []

    while True:  # Screen update
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                pygame.quit()
                save_to_excel(data)
                sys.exit()
            if event.type == pygame.MOUSEBUTTONDOWN:
                mouse_down = True
            if event.type == pygame.MOUSEBUTTONUP:
                mouse_down = False

        # If the mouse is pressed, add circles continuously
        if mouse_down:
            mouse_pos = pygame.mouse.get_pos()
            circles.append(create_circle(space, mouse_pos))

        # Apply aerodynamics
        apply_aerodynamics(circles)

        # Remove circles that have run off the screen
        circles = [circle for circle in circles if circle.body.position.x < 1200]

        # Data collection every 5 seconds
        current_time = time.time()
        elapsed_time = current_time - start_time
        if current_time - last_data_collection_time >= data_collection_interval:
            average_speed = calculate_average_speed(circles)
            data.append((int(elapsed_time), average_speed))
            last_data_collection_time = current_time

        screen.fill((217, 217, 217))  # Background color
        draw_circles(circles, screen, circle_surface)
        draw_shapes(shapes, screen)
        average_speed = calculate_average_speed(circles)
        display_average_speed(average_speed, screen, font)
        space.step(1 / 50)  # 0.02
        pygame.display.update()
        clock.tick(120)

if __name__ == "__main__":
    main()
